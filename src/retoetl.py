import pandas as pd
import os
import re
from word2number import w2n
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Función para estandarizar los valores alfabéticos a numéricos
def palabras_a_numeros(texto):
    try:
        texto = texto.lower().strip()
        if texto == "cero":
            return 0
        return w2n.word_to_num(texto)
    except ValueError:
        return texto  # Retorna sin cambios si no es procesable

# Función para configurar en modo de tabla los datos procesados
def configurar_tabla_excel(output_file, nombre_tabla="DatosProcesados"):
    wb = load_workbook(output_file)
    ws = wb.active

    # Definir el rango de la tabla
    num_filas = ws.max_row
    num_columnas = ws.max_column
    rango = f"A1:{chr(64 + num_columnas)}{num_filas}"

    # Configurar la tabla
    tabla = Table(displayName=nombre_tabla, ref=rango)
    estilo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)
    wb.save(output_file)

def process_csv_to_excel():
    # Obtiene la ruta del escritorio del usuario
    desktop = os.path.join(os.path.expanduser("~"), r"Desktop\Tec MTY\data")

    # Define la ruta de entrada y salida
    input_file = os.path.join(desktop, "BD_OPORTUNIDADES_23_24.csv")
    output_file = os.path.join(desktop, "OPORTUNIDADES_PROCESADO.xlsx")

    try:
        # Lee el archivo CSV
        data = pd.read_csv(input_file)

        # Identifica las columnas
        columnas = data.columns
        print(f"Columnas detectadas: {columnas}")

        # Limpieza de datos
        data = data.drop_duplicates() 
        data = data.dropna(subset=['Zona']) 

        # Transformar a mayúsculas las columnas de IDs
        for col in ['IdOportunidad', 'IdEmpresa', 'IdPropietario']:
            if col in data.columns:
                data[col] = data[col].astype(str).str.upper()

        # Transformación de datos
        if 'Importe' in data.columns:
            data['Importe'] = data['Importe'].apply(
                lambda x: palabras_a_numeros(x) if isinstance(x, str) else x
            )
            data['Importe'] = pd.to_numeric(data['Importe'], errors='coerce')  # Asegura valores numéricos

        if 'FechaCierre' in data.columns:
            data['FechaCierre'] = pd.to_datetime(data['FechaCierre'], errors='coerce', dayfirst=True)
            data['Año de Cierre'] = data['FechaCierre'].dt.year  # Extraer año
            data['Mes de Cierre'] = data['FechaCierre'].dt.month  # Extraer mes
            data['FechaCierre'] = data['FechaCierre'].dt.strftime('%d/%m/%Y')

        # Filtrar filas donde Importe es 0 o no existe valor
        if 'Importe' in data.columns:
            data = data[~(data['Importe'].isna() | (data['Importe'] == 0))]

        # Nuevas columnas calculadas para el dashboard
        if 'Importe' in data.columns:
            data['Rango Importe'] = pd.cut(data['Importe'], bins=[0, 10000, 50000, 100000, float('inf')],
                                           labels=['Bajo', 'Medio', 'Alto', 'Muy Alto'])

        # Nueva columna: Estado Participantes
        if 'Participantes' in data.columns:
            data['Estado Participantes'] = data['Participantes'].apply(
                lambda x: 'Activo' if pd.to_numeric(x, errors='coerce') > 0 else 'Inactivo' if pd.to_numeric(x, errors='coerce') == 0 else 'Desconocido'
            )
        else:
            data['Estado Participantes'] = 'Desconocido'

        # Verifica transformaciones
        print("Transformaciones completadas.")

        # Guarda los datos en un archivo Excel
        data.to_excel(output_file, index=False, engine='openpyxl')

        # Configura el formato de tabla en Excel
        configurar_tabla_excel(output_file)

        print(f"Archivo procesado y guardado en: {output_file}")
    except FileNotFoundError:
        print(f"El archivo '{input_file}' no se encontró en la ruta especificada.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

# Ejecuta la función
if __name__ == "__main__":
    process_csv_to_excel()
